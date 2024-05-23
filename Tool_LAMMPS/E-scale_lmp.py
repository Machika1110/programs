import os
import re
import math
import openpyxl as px
from scale_lmp import scaler
import matplotlib.pyplot as plt

# check str or int
def check_int(text):
    if text.isdigit():
        return int(text)
    else:
        return text

# for sort list of directory
def natural_keys(dir_list):
    tmp = re.split(r'(\d+)', dir_list)
    result = []
    for i in tmp:
        result.append(check_int(i))
    return result

# get directory
path = os.getcwd()
dirlist = os.listdir(path)
dir_list = []
for i in dirlist:
    if os.path.isdir(i):
        dir_list.append(i)
        
# sort directory
dir_list = sorted(dir_list, key=natural_keys)

volume = []
index_num = []
energy = []

# get data
for dir in dir_list:
    if os.path.isfile(dir+"/log.lammps"):
        with open(dir+"/log.lammps", "r") as f:
            line = f.readlines()
            for i in range(len(line)):
                if re.search("thermo_style", line[i]):
                    style = line[i].split()
                    for j in range(len(style)):
                        if style[j] == "etotal":
                            index_etotal = j-2
                
                if re.search("Step", i):
                    energy_tmp = line[i+1].split()[index_etotal]
                    energy.append()

# make graphs    
wb = px.Workbook()
ws = wb.worksheets[0]
ws.title = "E-scale_graph"
ws.cell(2, 2).value = "scaling"
ws.cell(2, 3).value = "energy"

for i in range(len(volume)):
    ws.cell(3+i, 2).value = float(scaler[i])
    ws.cell(3+i, 3).value = float(energy[i])

wb.save("E-scale.xlsx")

# make graph by matplotlib
# Create a plot
plt.figure(figsize=(8, 6))

# Plot the data
plt.plot(scaler, energy, marker='o', linestyle='-', color='b', label='Energy vs. Scaler')

# Add labels and title
plt.xlabel('Scaler')
plt.ylabel('Energy')
plt.title('Energy vs. Scaler')

# Add a legend
plt.legend()

# Add a grid
plt.grid(True)

# Save the plot to a file
plt.savefig('E-scale.png')

# Show the plot
plt.show()
