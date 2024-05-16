# make graphs of log.lammps based "thermo_style custom" included "step" in LEFTMOST !!!
# this program can't apply thermo_style except "custom" !!

import os
import re
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook

extracted_lines = []
thermo_block = 0


# get all information based on thermo_style
if os.path.exists("log.lammps"):
    with open("log.lammps", "r") as f:
        lines = f.readlines()
        for i in lines:
            if re.search("thermo_style", i) and not re.search("WARNING", i):
                #print(i)
                i = i.split()
                #print(i)
                tag = i[2:]
                #print(tag)
                if tag[0] != "step":
                    print("thermo_style has no step information.\n")
                    print("Program can't be executed !!\n")
                    exit()
            
        start_index = None
        end_index = None

            # search for lines
        for i, line in enumerate(lines):
            if start_index is None and re.search(r'^Step', line):
                start_index = i + 1
                #print(start_index)
            elif start_index is not None and re.search(r'^Loop', line):
                end_index = i
                #print(end_index)

            if start_index is not None and end_index is not None:
                extracted_lines.append(lines[start_index:end_index])
                thermo_block += 1
                start_index = None
                end_index = None

        for i in range(thermo_block):
            for j in range(len(extracted_lines[i])):
                extracted_lines[i][j] = extracted_lines[i][j].split()


# combine "tag" + "extracted_lines" to one matrix
extracted_lines.insert(0, tag)
#print(extracted_lines)
matrix = np.vstack(extracted_lines)
#print(matrix)


# make graphs with matplotlib
if not os.path.exists("trace_lmp_graphs"):
    os.makedirs("trace_lmp_graphs")
# x-axis data (step)
x_data = matrix[1:, 0].astype(int)

# plot data for each column and save
for col_idx in range(1, matrix.shape[1]):
    # get y-axis data
    y_data = matrix[1:, col_idx].astype(float)
    
    # plot graph
    plt.plot(x_data, y_data, marker='o', linestyle='-')

    # set title and labels
    plt.title(f'{matrix[0, col_idx]} vs Step')
    plt.xlabel('Step')
    plt.ylabel(matrix[0, col_idx])

    # display grid
    plt.grid(True)

    # save graph
    plt.savefig(f"trace_lmp_graphs/{matrix[0, col_idx]}.png")  # or specify 'jpeg' or other file formats

    # clear plot for next graph
    plt.clf()

# close plot display
plt.close()


# create new Excel workbook
wb = Workbook()
ws = wb.active

# write data to Excel
for row_idx, row_data in enumerate(matrix):
    for col_idx, cell_value in enumerate(row_data):
        if row_idx == 0:
            ws.cell(row=row_idx+1, column=col_idx+1, value=str(cell_value))
        else:
            ws.cell(row=row_idx+1, column=col_idx+1, value=float(cell_value))

# save Excel file
wb.save("trace_lmp_graphs/trace_lmp.xlsx")
